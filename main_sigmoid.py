import random
import torch.nn.functional as F
import torch
from datasets import load_metric
from openpyxl import load_workbook
from prettytable import PrettyTable
from torch import nn
from torch.utils.data import TensorDataset, DataLoader
from transformers import AutoTokenizer, AutoConfig, get_linear_schedule_with_warmup, AdamW, BertModel
from sklearn.metrics import classification_report
from total_extra import extra
from openpyxl import load_workbook, Workbook
import pandas as pd

def data_loader_file(file):
    wb = load_workbook(file)
    sheets = wb.worksheets
    sheet1 = sheets[0]
    rows = sheet1.rows
    id_text = {}
    num = 0
    cancer = {0: '腺癌', 1: '肺良性疾病', 2: '鳞癌', 3: '无法判断组织分型', 4: '复合型', 5: '转移癌', 6: '小细胞癌', 7: '大细胞癌'}
    t = {0: '无', 1: '转移'}
    l_t = {0: '无', 1: '淋巴转移'}
    id_cancer = {j:i for i,j in cancer.items()}
    id_t = {j:i for i,j in t.items()}
    id_l_t = {j:i for i,j in l_t.items()}
    for row in rows:
        if num == 0:
            num+=1
            continue
        row_val = [col.value for col in row]
        if len(row_val)<8:
            continue
        # print(row_val)
        id_text[row_val[0]] = {"text":"","list":['wu',0,0,0,'wu'],"size":[],"size_index":[]}
        id_text[row_val[0]]["text"] = row_val[1]
        id_text[row_val[0]]["list"][1] = id_cancer[row_val[3]]
        id_text[row_val[0]]["list"][2] = id_t[row_val[4]]
        id_text[row_val[0]]["list"][3] = id_l_t[row_val[5]]
        if len(row_val)>9 and row_val[9]!=None:
            id_text[row_val[0]]["list"][1] = id_cancer[row_val[9]]
        if row_val[7]==None:
            id_text[row_val[0]]["size"] = []
        else:
            aa = str(row_val[7]).split(' ')
            id_text[row_val[0]]["size"] = aa
        if row_val[8]==None:
            continue
        k = row_val[8]
        kk = k.split('\t')
        size_index_t = []
        for kkk in kk:
            s = kkk.split(' ')
            s_0 = int(s[0])
            s_1 = int(s[1])
            size_index_t.append([s_0,s_1])

        id_text[row_val[0]]["size_index"] = size_index_t
    return id_text



def mark_data_loader(file):
    wb = load_workbook(file)
    sheets = wb.worksheets
    sheet1 = sheets[0]
    rows = sheet1.rows
    id_text = {}
    num = 0
    for row in rows:
        if num == 0:
            num+=1
            continue
        row_val = [col.value for col in row]
        # print(row_val)
        id_text[row_val[1]] = row_val[5:10]
    # print(id_text)
    return id_text

def train_test(data):
    transfer_id = []
    exist_id = []
    train_id = []
    test_id = []
    for i,j in data.items():
        if j["list"][2]==1:
            transfer_id.append(i)
    # print(transfer_id)
    train_id+= transfer_id[:int(len(transfer_id)*0.8)]
    test_id+= transfer_id[int(len(transfer_id)*0.8):]
    # print(train_id,test_id)
    exist_id+= transfer_id
    cancer_id_list = {2:[],4:[],5:[],6:[],7:[]}
    for i,j in data.items():
        if i not in exist_id and j["list"][1] in [2,4,5,6,7]:
            cancer_id_list[j["list"][1]].append(i)
            exist_id.append(i)
    # print(cancer_id_list)
    # print([len(i) for _,i in cancer_id_list.items()])
    for i,j in cancer_id_list.items():
        train_id+= j[:int(len(j)*0.8)]
        test_id+= j[int(len(j) * 0.8):]

    other_id = []
    for i,j in data.items():
        if i not in exist_id:
            if j["list"][1]==0 and random.randint(0,7)!=0:
                continue
            other_id.append(i)
    # print(other_id)
    random.shuffle(other_id)
    # print(other_id)

    train_id += other_id[:int(len(other_id)*0.8)]
    test_id += other_id[int(len(other_id) * 0.8):]
    # can_id = {i:1 for i in range(8)}
    # for i in test_id:
    #     can_id[data[i]["list"][1]] += 1
    # print(can_id)

    return train_id,test_id

def data_loader(data):
    data_keys = data.keys()
    text_input = torch.zeros((len(data_keys),512)).long()
    mask_input = torch.zeros((len(data_keys),512), dtype=torch.uint8)
    seq_out = torch.zeros((len(data_keys),512)).long()
    cancer_label = torch.zeros(len(data_keys)).long()
    t_l = torch.zeros(len(data_keys)).long()
    l_t_l = torch.zeros(len(data_keys)).long()
    text_dd = []
    cancar_list = [0 for i in range(8)]
    t_l_list = [0,0]
    l_t_l_list = [0,0]
    data_t = [j for _,j in data.items()]
    for index,i in enumerate(data_t):
        one = i
        # print(one)
        text_dd.append(one["text"][:510])
        # print(one)
        for j in one["size_index"]:
            seq_out[index][j[0]+1] = 1
            seq_out[index][j[0]+2:j[1]+1] = 1
            # print(j)
            # print(seq_out[index])
        text = tokenizer.convert_tokens_to_ids(['[CLS]']+list(one["text"][:510])+['[SEP]'])
        text_input[index][:len(text)] = torch.tensor(text)
        mask_input[index][:len(text)] = 1
        cancer_label[index] = one["list"][1]
        cancar_list[one["list"][1]] += 1
        t_l[index] = one["list"][2]
        t_l_list[one["list"][2]] += 1
        l_t_l[index] = one["list"][3]
        l_t_l_list[one["list"][3]] += 1
    print("病理组织",cancar_list)
    print("转移",t_l_list)
    print("癌转移",l_t_l_list)
    # for i in range(5):
    #     one = data[id[i]]
    #     print(one["text"])
    #     print(text_input[i])
    #     print(seq_out[i])
    return TensorDataset(text_input, seq_out, mask_input, cancer_label, t_l, l_t_l),text_dd
# def data_loader(id,data):
#     text_input = torch.zeros((len(id),512)).long()
#     mask_input = torch.zeros((len(id),512), dtype=torch.uint8)
#     seq_out = torch.zeros((len(id),512)).long()
#     cancer_label = torch.zeros(len(id)).long()
#     t_l = torch.zeros(len(id)).long()
#     l_t_l = torch.zeros(len(id)).long()
#     text_dd = []
#     cancar_list = [0 for i in range(8)]
#     t_l_list = [0,0]
#     l_t_l_list = [0,0]
#     for index,i in enumerate(id):
#         one = data[i]
#         text_dd.append(one["text"][:510])
#         # print(one)
#         for j in one["size_index"]:
#             seq_out[index][j[0]+1] = 1
#             seq_out[index][j[0]+2:j[1]+1] = 1
#             # print(j)
#             # print(seq_out[index])
#         text = tokenizer.convert_tokens_to_ids(['[CLS]']+list(one["text"][:510])+['[SEP]'])
#         text_input[index][:len(text)] = torch.tensor(text)
#         mask_input[index][:len(text)] = 1
#         cancer_label[index] = one["list"][1]
#         cancar_list[one["list"][1]] += 1
#         t_l[index] = one["list"][2]
#         t_l_list[one["list"][2]] += 1
#         l_t_l[index] = one["list"][3]
#         l_t_l_list[one["list"][3]] += 1
#     print("病理组织",cancar_list)
#     print("转移",t_l_list)
#     print("癌转移",l_t_l_list)
#     # for i in range(5):
#     #     one = data[id[i]]
#     #     print(one["text"])
#     #     print(text_input[i])
#     #     print(seq_out[i])
#     return TensorDataset(text_input, seq_out, mask_input, cancer_label, t_l, l_t_l),text_dd

class Word_BERT(nn.Module):
    def __init__(self, seq_label=1,cancer_label=8,transfer_label=2,ly_transfer=2):
        super(Word_BERT, self).__init__()
        self.bert = BertModel.from_pretrained('/home/kelab/Documents/transformers_/bert-base-zh')
        # self.bert_config = self.bert.config
        self.out = nn.Sequential(
            # nn.Linear(768,256),
            # nn.ReLU(),
            nn.Dropout(0.1),
            nn.Linear(768, seq_label)
        )
        self.cancer = nn.Sequential(
            nn.Dropout(0.1),
            nn.Linear(768, cancer_label)
        )
        self.transfer = nn.Sequential(
            nn.Dropout(0.1),
            nn.Linear(768, transfer_label)
        )
        self.ly_transfer = nn.Sequential(
            nn.Dropout(0.1),
            nn.Linear(768, ly_transfer)
        )

    def forward(self, word_input, masks):
        # print(word_input.size())
        output = self.bert(word_input, attention_mask=masks)
        sequence_output = output.last_hidden_state
        pool = output.pooler_output
        # print(sequence_output.size())
        # print(pool.size())
        out = self.out(sequence_output)
        cancer = self.cancer(pool)
        transfer = self.transfer(pool)
        ly_transfer = self.ly_transfer(pool)
        return out,cancer,transfer,ly_transfer

def train_few(train_data, test_data,cancer_id_l,transfer_id_l,lymph_transfer_id_l,text):
    train_batch_size = 30
    test_batch_size = 64
    train_iter = DataLoader(train_data, shuffle=True, batch_size=train_batch_size)
    test_iter = DataLoader(test_data, shuffle=False, batch_size=test_batch_size)

    label_vocab = {"O":0,"B-size":1,"I-size":2}
    id_aux_label = {j: i for i, j in label_vocab.items()}
    model = Word_BERT()
    model.to(torch.device('cuda:1'))

    param_optimizer = list(model.named_parameters())
    no_decay = ['bias', 'LayerNorm.weight']
    no_bert = ['bert']
    optimizer_grouped_parameters = [
        {'params': [p for n, p in param_optimizer if
                    ((not any(nd in n for nd in no_decay)) and any(nd in n for nd in no_bert))], 'weight_decay': 0.01,
         'lr': 8e-5},
        {'params': [p for n, p in param_optimizer if
                    ((any(nd in n for nd in no_decay)) and any(nd in n for nd in no_bert))], 'weight_decay': 0.0,
         'lr': 8e-5},
        {'params': [p for n, p in param_optimizer if
                    ((not any(nd in n for nd in no_decay)) and (not any(nd in n for nd in no_bert)))],
         'weight_decay': 0.01, 'lr': 1e-2},
        {'params': [p for n, p in param_optimizer if
                    ((any(nd in n for nd in no_decay)) and (not any(nd in n for nd in no_bert)))], 'weight_decay': 0.0,
         'lr': 1e-2}
    ]

    optimizer = AdamW(optimizer_grouped_parameters, correct_bias=True)
    warm_ratio = 0.1
    epochs = 100
    print("train_batch_size", train_batch_size)
    print("graident_steps", 1)
    # print(len(train_data))
    total_steps = (len(train_data) // train_batch_size + 1) * epochs

    print("total_steps", total_steps)
    scheduler = get_linear_schedule_with_warmup(optimizer, num_warmup_steps=warm_ratio * total_steps,
                                                num_training_steps=total_steps)
    metric = load_metric("./seqeval_metric.py")
    # id_aux_label = {0: "B", 1: "O"}
    loss_fun = torch.nn.BCEWithLogitsLoss(reduction='none')
    model.train()
    max_f1 = 0
    for epoch in range(epochs):
        for step, batch in enumerate(train_iter):
            text_input, seq_out, mask_input, cancer_label, t_l, l_t_l = batch
            text_input, seq_out, mask_input, cancer_label, t_l, l_t_l = text_input.to(torch.device('cuda:1')), seq_out.to(torch.device('cuda:1')), mask_input.to(torch.device('cuda:1')), cancer_label.to(torch.device('cuda:1')), t_l.to(torch.device('cuda:1')), l_t_l.to(torch.device('cuda:1'))
            out,cancer,transfer,ly_transfer = model(text_input, mask_input)
            # print(out.view(-1),seq_out.view(-1))
            loss = loss_fun(out.view(-1),seq_out.float().view(-1))
            seq_loss = torch.sum(loss*mask_input.view(-1))/torch.sum(mask_input)
            # seq_loss = F.cross_entropy(out.view(-1,3),seq_out.view(-1))
            cancer_loss = F.cross_entropy(cancer.view(-1,8),cancer_label.view(-1))
            t_loss = F.cross_entropy(transfer.view(-1,2),t_l.view(-1))
            ly_loss = F.cross_entropy(ly_transfer.view(-1,2),l_t_l.view(-1))

            loss = seq_loss+cancer_loss+t_loss+ly_loss
            loss.backward()
            # loss.backward()
            nn.utils.clip_grad_value_(model.parameters(), 1)
        if (step+1)%2==0 or (step+1)==len(train_iter):
            optimizer.step()
            scheduler.step()
            model.zero_grad()
        print("loss",loss.item())
        if (epoch+1)>=40 and (epoch+1)%5==0:
        # if (epoch + 1) == epochs:
            model.eval()
            with torch.no_grad():
                out_l,cancer_l,transfer_l,ly_transfer_l = [[],[]],[[],[]],[[],[]],[[],[]]
                for batch in test_iter:
                    text_input, seq_out, mask_input, cancer_label, t_l, l_t_l = batch
                    text_input, seq_out, mask_input, cancer_label, t_l, l_t_l = text_input.to(torch.device('cuda:1')), seq_out.to(torch.device('cuda:1')), mask_input.to(torch.device('cuda:1')), cancer_label.to(torch.device('cuda:1')), t_l.to(torch.device('cuda:1')), l_t_l.to(torch.device('cuda:1'))
                    out,cancer,transfer,ly_transfer = model(text_input, mask_input)
                    out = F.sigmoid(out).squeeze(2).cpu()
                    print(out.size())
                    out = out.numpy().tolist()
                    # print(out.size())
                    # out = out.argmax(dim=-1).cpu().numpy().tolist()
                    cancer = cancer.argmax(dim=-1).cpu().numpy().tolist()
                    transfer = transfer.argmax(dim=-1).cpu().numpy().tolist()
                    ly_transfer = ly_transfer.argmax(dim=-1).cpu().numpy().tolist()
                    seq_out, cancer_label, t_l, l_t_l = seq_out.cpu().numpy().tolist(), cancer_label.cpu().numpy().tolist(), t_l.cpu().numpy().tolist(), l_t_l.cpu().numpy().tolist()
                    for i in range(text_input.size()[0]):
                        out_l[0].append(out[i][1:mask_input[i].sum()-1])
                        out_l[1].append(seq_out[i][1:mask_input[i].sum()-1])
                        cancer_l[0].append(cancer[i])
                        cancer_l[1].append(cancer_label[i])
                        transfer_l[0].append(transfer[i])
                        transfer_l[1].append(t_l[i])
                        ly_transfer_l[0].append(ly_transfer[i])
                        ly_transfer_l[1].append(l_t_l[i])
            
            # with open('./text.txt',mode='w',encoding='utf-8') as f:
            #     for ii in range(len(mer_preds)):
            #         for jj in range(len(mer_preds[ii])):
            #             # print(text[ii][jj],mer_preds[ii][jj],mer_labels[ii][jj])
            #             f.writelines(text[ii][jj]+'\t'+mer_preds[ii][jj]+'\t'+mer_labels[ii][jj]+'\n')
            #         f.writelines('\n')
            cancer_result = classification_report(cancer_l[1], cancer_l[0], target_names=cancer_id_l,output_dict=True)
            print(cancer_result)
            if cancer_result["macro avg"]['f1-score']>max_f1:
                torch.save(model.state_dict(), "./result/16/model.pth")
                with open('./result/16/epoch.txt', mode='a', encoding='utf-8') as f:
                    f.writelines(str(epoch)+'\n')
                max_f1 = cancer_result["macro avg"]['f1-score']
                df = pd.DataFrame(cancer_result).transpose()
                df.to_csv("./result/16/cancer_result.csv", index= True)
                transfer_result = classification_report(transfer_l[1], transfer_l[0], target_names=transfer_id_l,output_dict=True)
                df = pd.DataFrame(transfer_result).transpose()
                df.to_csv("./result/16/transfer_result.csv", index= True)
                transfer_l_result = classification_report(ly_transfer_l[1], ly_transfer_l[0], target_names=lymph_transfer_id_l,output_dict=True)
                df = pd.DataFrame(transfer_l_result).transpose()
                df.to_csv("./result/16/transfer_l_result.csv", index= True)
                for thresold in [0.05,0.08,0.1,0.15,0.2,0.3,0.4]:
                    pred_thresold = [[1 if jj>thresold else 0  for jj in ii] for ii in out_l[0]]
                    for ii in range(len(pred_thresold)):
                        for jj,kk in enumerate(pred_thresold[ii]):
                            if kk==1 and (pred_thresold[ii][jj-1]==1 or pred_thresold[ii][jj-1]==2):
                                pred_thresold[ii][jj] = 2
                    for ii in range(len(out_l[1])):
                        for jj,kk in enumerate(out_l[1][ii]):
                            if kk==1 and (out_l[1][ii][jj-1]==1 or out_l[1][ii][jj-1]==2):
                                out_l[1][ii][jj] = 2
                    
                    mer_preds = [[id_aux_label[j] for j in i] for i in pred_thresold]
                    mer_labels = [[id_aux_label[j] for j in i] for i in out_l[1]]
                    # print(mer_preds[:5])
                    # print(mer_labels[:5])
                    metric.add_batch(
                        predictions=mer_preds,
                        references=mer_labels,
                    )
                    results = metric.compute()
                    f1 = results["overall_f1"]
                    p = results["overall_precision"]
                    r = results["overall_recall"]
                    with open('./result/16/size.txt', mode='a', encoding='utf-8') as f:
                        f.writelines("阈值"+str(thresold)+'\n')
                        f.writelines("F1:" + str(round(f1 * 100, 2)) + "pre" + str(round(p * 100, 2)) + "recall" + str(round(r * 100, 2)) + "\n")
                    # print("阈值",str(thresold))
                # print(results)
            model.train()
        
            


if __name__=='__main__':
    raw_mark_ali = extra()
    # for i,j in raw_mark_ali.items():
    #     print(i,j)

    cancer = {'腺癌': 0, '肺良性疾病': 1, '鳞癌': 2, '无法判断组织分型': 3, '复合型': 4, '转移癌': 5, '小细胞癌': 6, '大细胞癌': 7}
    num = 0
    cancer_id = {}
    for i in cancer.keys():
        cancer_id[i] = num
        num += 1
    transfer_id = {'无': 0, '转移': 1}
    lymph_transfer_id = {'无': 0, '淋巴转移':1}
    print(cancer_id)
    print(transfer_id)
    print(lymph_transfer_id)
    cancer_id_l = [i for i,_ in cancer_id.items()]
    transfer_id_l = [i for i,_ in transfer_id.items()]
    lymph_transfer_id_l = [i for i,_ in lymph_transfer_id.items()]

    print("总数",len(raw_mark_ali))
    # raw_mark_convert = {}
    cancer_nums = {'腺癌': 0, '肺良性疾病': 0, '鳞癌': 0, '无法判断组织分型': 0, '复合型': 0, '转移癌': 0, '小细胞癌': 0, '大细胞癌': 0}
    transfer_nums = {'无': 0, '转移': 0}
    lymph_transfer_nums = {'无': 0, '淋巴转移':1}
    for i,j in raw_mark_ali.items():
        # print(raw_mark_ali[i]["list"])
        cancer_nums[raw_mark_ali[i]["list"][0]] += 1
        transfer_nums[raw_mark_ali[i]["list"][1]] += 1
        lymph_transfer_nums[raw_mark_ali[i]["list"][2]] += 1
        raw_mark_ali[i]["list"] = ['wu']+raw_mark_ali[i]["list"]+['wu']
        raw_mark_ali[i]["list"][1] = cancer_id[raw_mark_ali[i]["list"][1]]
        raw_mark_ali[i]["list"][2] = transfer_id[raw_mark_ali[i]["list"][2]]
        raw_mark_ali[i]["list"][3] = lymph_transfer_id[raw_mark_ali[i]["list"][3]]
    # for i,j in raw_mark_ali.items():
    #     print(i,j)
    # print(raw_mark_ali)
    print(cancer_nums,transfer_nums,lymph_transfer_nums)
    train_id,test_id = train_test(raw_mark_ali)
    id_cancer = {j:i for i,j in cancer_id.items()}
    id_trans = {j:i for i,j in transfer_id.items()}
    id_l_trans = {j:i for i,j in lymph_transfer_id.items()}

    print(id_cancer)
    print(id_trans)
    print(id_l_trans)
    

    tokenizer = AutoTokenizer.from_pretrained('/home/kelab/Documents/transformers_/bert-base-zh', use_fast=True)

    config = AutoConfig.from_pretrained('/home/kelab/Documents/transformers_/bert-base-zh')
    # vocab = tokenizer.get_vocab()
    # print(vocab)
    # train_data,text = data_loader(train_id,raw_mark_ali)
    # test_data,text = data_loader(test_id,raw_mark_ali)
    
    
    train_text = data_loader_file('./训练集_排错.xlsx')
    test_text = data_loader_file('./测试集_排错.xlsx')
    print(len(train_text.keys()))
    for i,j in raw_mark_ali.items():
        # #第一版
        # if int(i) not in train_text.keys() and int(i) not in test_text.keys():
        #     train_text[i] = j
        #第二版
        if int(i) not in train_text.keys() and int(i) not in test_text.keys() and (j["list"][1]==0 and random.randint(0,40)==0):
            train_text[i] = j

    train_data,text = data_loader(train_text)
    test_data,text = data_loader(test_text)
    print(len(train_text.keys()))

    train_few(train_data,test_data,cancer_id_l,transfer_id_l,lymph_transfer_id_l,text)

    


